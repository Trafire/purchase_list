from autof2.dailytasks import purchaselist

from openpyxl import Workbook
from openpyxl.worksheet import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl import load_workbook
import openpyxl

import datetime
from datetime import date
from autof2.interface import send_data
from autof2.navigation import navigation
import os
import os.path


def get_date_sunday(year, week):
    d = str(year) + '-W' + str(week)
    return datetime.datetime.strptime(d + '-1', "%Y-W%W-%w") - datetime.timedelta(days=1)


def get_today():
    year = datetime.date.today().strftime("%Y")
    week = datetime.date.today().strftime("%W")
    day = datetime.date.today().strftime("%w")
    d = str(year) + '-W' + str(week)
    return datetime.datetime.strptime(d + '-' + day, "%Y-W%W-%w")


def get_current_week(add=0):
    date = (datetime.datetime.now() + datetime.timedelta(days=add)).isocalendar()
    if date[2] == 7:
        d = datetime.datetime.now() + datetime.timedelta(days=7)
        date = d.isocalendar()
    week = date[1]
    year = date[0]
    return (year, week)


def get_order_week(year, week):
    current = get_date_sunday(year, week)
    product = []
    print("\nstarting Week %i:" % week)
    for i in range(7):
        str_date = current.strftime('%d/%m/%y')
        if current >= get_today():
            print("\tprocessing day - %s" % current.strftime('%d/%m/%y'), end=" ")
            new_product = purchaselist.run_all_purchase_list_report(str_date, str_date)

            for p in new_product:
                p.date = str_date
            print(" lines found = %i" % len(new_product))
            product.extend(new_product)
            ##            print(current.strftime('%d/%m/%y'))

            send = send_data.SendData()
            send.send('{LEFT}')
        current += datetime.timedelta(days=1)
    print("Week %i total lines = %i" % (week, len(product)))
    return product


def make_order_sheet(wb, product, year, week):
    ws = wb.active
    rows = 1
    ws.append(product[0].excel_heading())
    for line in product:
        ws.append(line.excel_data())
        rows += 1

    right_corner = chr(64 + len(product[0].excel_heading())) + str(rows)

    # define a table style
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                          showRowStripes=True)
    # create a table
    table = openpyxl.worksheet.table.Table(ref='A1:' + right_corner,
                                           displayName='orders',
                                           tableStyleInfo=mediumStyle)
    # add the table to the worksheet
    ws.title = "orders"
    ws.add_table(table)


def go_to_puchase_list():
    for i in range(10):
        print(i)
        if navigation.to_purchase_list():
            return True
    return False


###
######import time
######
######import sys
######import traceback
######
######class TracePrints(object):
######  def __init__(self):
######    self.stdout = sys.stdout
######  def write(self, s):
######    self.stdout.write("Writing %r\n" % s)
######    traceback.print_stack(file=self.stdout)
######    time.sleep(10)
######
######sys.stdout = TracePrints()s
#


future_weeks = 1
print("stuff")
if go_to_puchase_list():
    for i in range(future_weeks + 1):
        date = get_current_week()
        year = date[0]
        week = date[1]
        week += i
        product = get_order_week(year, week)
        if product:
            wb = Workbook()
            make_order_sheet(wb, product, year, week)
            # create directory
            directory = os.getcwd() + 'orders\\%s\\week %s' % (year, week)
            if not os.path.exists(directory):
                os.makedirs(directory)
            # save the workbook file
            filename = directory + '\\week ' + str(week) + ' orders' + '.xlsx'
            ##        filename = "test2.xlsx"

            bought = {}

            if os.path.isfile(filename):
                wb2 = load_workbook(filename)
                a = (wb2.get_sheet_by_name('purchases'))
                index = 0

                for row in a.rows:
                    if index == 0:
                        index += 1
                        categories_order = row
                    else:
                        p = {}
                        for i in range(len(categories_order)):
                            p[categories_order[i].value] = row[i].value
                        if p['PurchaseID'] in bought:

                            bought[p['PurchaseID']]['Confirmed'] += p['Confirmed']
                        else:
                            bought[p['PurchaseID']] = p
                        bought[p['PurchaseID']]['Ordered'] = 0

            for p in product:
                if p.key not in bought:
                    bought[p.key] = p.excel_order_dict_vers()
                else:
                    bought[p.key]['Ordered'] += p.quantity

            product_list = []

            for b in bought:
                product_list.append(bought[b])

            ##    wb = Workbook()
            ws2 = wb.create_sheet()
            rows = 1
            headings = (
            "PurchaseID", "f2_supplier", "Category", "Variety", "Colour", "Grade", "Supplier", "Price", "Ordered",
            "Confirmed")
            ws2.append(headings + ("Total",))
            for line in bought:
                l = []
                for h in headings:
                    l.append(bought[line][h])
                l.append("=J%s - I%s" % (rows + 1, rows + 1))
                ws2.append(l)
                rows += 1

            right_corner = chr(64 + 1 + len(product[0].excel_order_headings())) + str(rows)

            # define a table style
            mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                                  showRowStripes=True)
            # create a table
            table = openpyxl.worksheet.table.Table(ref='A1:' + right_corner,
                                                   displayName='purchases',
                                                   tableStyleInfo=mediumStyle)
            # add the table to the worksheet
            ws2.title = "purchases"
            ws2.add_table(table)

            # save the workbook file
            ##    wb.save('test_1'.replace(':','-').replace('.','-') + '.xlsx')
            ##
            wb.save(filename)

        #os.startfile(filename)
